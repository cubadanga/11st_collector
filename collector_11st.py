import re
import os
import sys
import time
import random
import requests
import pandas as pd
import numpy as np
import configparser
from openpyxl import load_workbook, Workbook

from colorama import init, Fore
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
import urllib
from urllib.parse import parse_qs, urlparse
from urllib.error import HTTPError
import urllib.request
from urllib.request import urlopen
from urllib.parse import urlparse, parse_qs

# 크롬드라이버 자동업데이트
tday = time.time()
tday_s = time.strftime('%Y%m%d-%H%M%S',time.localtime(time.time()))
tday_v = time.strftime('%Y/%m/%d-%H:%M', time.localtime(time.time()))



def loadPassword(): #우선 'set.ini' 파일에 저장된 패스워드와 웹에 있는 패스워드가 일치하는지 확인한다.
    basedir = os.getcwd()
    ini_dir = os.path.join(basedir,'set.ini')

    # pc set.ini 파일의 저장된 pass워드 읽어오기
    properties = configparser.ConfigParser()
    properties.read(ini_dir)
    
    if 'DEFAULT' in properties and 'userpass' in properties['DEFAULT']:
        password = properties['DEFAULT']['userpass']
        return password
    else:
        print(Fore.RED + "오류 - 'userpass' key not found in set.ini file."+'\n')
        print(Fore.RESET + "엔터를 누르면 종료합니다.")
        aInput = input("")
        sys.exit()

def getPtag(url): # 웹페이지에 적어 놓은 password 텍스트를 크롤링해 추출하는 함수
    try:
        html = urlopen(url)
        
    except HTTPError as e:
        print(Fore.RED + '오류 - 네트워크오류 또는 패스워드url오류'+'\n')
        print(Fore.RESET + "엔터를 누르면 종료합니다.")
        aInput = input("")
        sys.exit()
    try:
        soup = BeautifulSoup(html,"html.parser")
        ptag = soup.find('p')
        
    except AttributeError as e:
        return None
    return ptag.text

def judge(password,passTag): #set.ini에 저장된 패스워드와 웹에 있는 패스워드를 비교하는 함수.
    if password == passTag:
        properties = configparser.ConfigParser()
        properties.set('DEFAULT','userpass',password)
        with open('./set.ini','w',encoding='utf-8') as F:
            properties.write(F)
                
        pass
    else:
        print(Fore.RED + "오류 - 저장된 패스워드가 없거나 올바른 패스워드가 아닙니다."+Fore.RESET+'\n')
        inputPass(password,passTag)

def inputPass(password,passTag): #패스워드가 틀렸을 때 콘솔에서 다시 입력을 받는 함수
    userPass = str(password)
    passTag = passTag
    print('\n' + "패스워드를 입력해 주세요.")
    userPass = input()
    judge(userPass, passTag)

def set_selenium():

    #브라우저 자동꺼짐 방지
    chrome_options = Options()
    chrome_options.add_experimental_option('detach',True)

    #헤더추가
    chrome_options.add_argument('--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/89.0.4389.82 Safari/537.36')
    chrome_options.add_argument(f'--disalbe-blink-fatures=AutomationControlled')
    chrome_options.add_argument("--headless=new")

    #불필요한 에러메시지 없애기
    chrome_options.add_experimental_option("excludeSwitches", ['enable-logging'])
    driver = webdriver.Chrome(options=chrome_options)
    return driver

def readExcel(urls_path): #셋팅 데이터와 유저가 입력한 제품 데이터를 읽어 온다.
    try:
        urls_DF = pd.read_excel(urls_path, sheet_name = 'url', header = 0)
        #setting_DF = pd.read_excel(setting_path, sheet_name = 'setting', header = 0)
        #setting_DF = setting_DF.fillna('')
        return urls_DF
        
    except ValueError as e:
        print(Fore.RED + '오류 - 엑셀 시트의 시트명이 다르거나 올바른 파일이 아닙니다.'+'\n')
        print(Fore.RESET + "엔터를 누르면 종료합니다.")
        aInput = input("")
        sys.exit()

    except FileNotFoundError as e:
        print(Fore.RED + '오류 - product.xlsx 파일을 찾을 수 없습니다.'+'\n'+'이런 경우, 파일명이 잘못된 경우가 대부분이었습니다.'+' 이 파일은 필수 파일입니다.'+'\n')
        print(Fore.RESET + "엔터를 누르면 종료합니다.")
        aInput = input("")
        sys.exit()

def createFolder(directory): # 결과 파일 저장 폴더 생성
    try:
        if not os.path.exists(directory):
            os.makedirs(directory)
    except OSError:
        print (Fore.RED + '오류 - Creating directory. ' +  directory)
        print(Fore.RESET + "엔터를 누르면 종료합니다.")
        aInput = input("")
        sys.exit()

#url 불러옴
def scrapping(url, driver):
    driver.get(url)
    time.sleep(0.5)
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
    html = driver.page_source
    html = html.encode('utf-8')
    soup = BeautifulSoup(html, 'html.parser')
    return soup

# 페이지에서 카테고리를 추출함.

def url_parser(soup):
    cat_items = soup.select('.pdp-category-list.list')
    if cat_items:
        last_cat_item = cat_items[-1]
        button_element = last_cat_item.select_one('button')
        if button_element.has_attr('link-url'):
            link_url = button_element['link-url']
        else:
            print("'link-url' attribute not found.")
            return ""
    else:
        print("카테고리를 찾지 못했습니다.")
        return ""

    parse_link = urlparse(link_url)
    query_params = parse_qs(parse_link.query)
    categori_num = ""

    if 'dispCtgrNo' in query_params:
        categori_num = query_params['dispCtgrNo'][0]
    return categori_num


def extractComp(soup, id_num, url_count, count_num):
    
    id_num = str(id_num)
    #categori_search = soup.select('.pdp-category-list list"')
    #print(f'실제 카테고리 ID : {categori_search}')

    img_main = soup.select('#productImg div.img_full img')  # 선택된 메인이미지 들을 추출하는데 1000x1000 위주로, 없으면 600x600
    main_img_list = []
    mainImgList = []
    
    for img in img_main:
        src = img['src']
        if src:
            jpg_urls = re.findall(r'https?://[^\s/$.?#].[^\s]*\.(?:jpg|png|jpeg|gif)', src)
            main_img_list.extend(jpg_urls)

    #가져온 이미지중 중복 파일명 제거. 1000x1000만 남김.
    for url in main_img_list:
        if '1000x1000' in url:
            mainImgList.append(url)
        elif '600x600' in url:
            mainImgList.append(url)

    if len(main_img_list) == 1:
        mainImage = 'main_'+id_num +'_'+'0.jpg' 
        addImage1 = '' 
        addImage2 = ''
        addImage3 = ''

    elif len(main_img_list) == 2:
        mainImage = 'main_'+id_num +'_'+'0.jpg'
        addImage1 = 'main_'+id_num +'_'+'1.jpg'
        addImage2 = ''
        addImage3 = ''
    
    elif len(main_img_list) == 3:   
        mainImage = 'main_'+id_num +'_'+'0.jpg'
        addImage1 = 'main_'+id_num +'_'+'1.jpg'
        addImage2 = 'main_'+id_num +'_'+'2.jpg'
        addImage3 = ''

    elif len(main_img_list) >= 4:
        mainImage = 'main_'+id_num +'_'+'0.jpg'
        addImage1 = 'main_'+id_num +'_'+'1.jpg'
        addImage2 = 'main_'+id_num +'_'+'2.jpg'
        addImage3 = 'main_'+id_num +'_'+'3.jpg'

    p_title = soup.select_one('title') #결과는 리스트 형태로 나옴.
    title = p_title.text #links 리스트에서 텍스트 요소를 가져옴.
    print(f'상품명 : {title}',file=f)
    print(f'상품명 : {title}')
    p_price = soup.select_one('.value') #href 속성 값만 가져옴
    txt_price = p_price.text
    txt_price = txt_price.replace('원', '')
    base_price = int(txt_price.replace(',', ''))

    #옵션의 추출
    try:
        li_tags = soup.select('.option_item_list li')

        result_list = []
        for li_tag in li_tags:
            optitemno = li_tag.get('data-optitemno')
            dtloptnm = li_tag.get('data-dtloptnm')
            txt_optPrice = li_tag.get('data-price')
            opt_price = int(txt_optPrice.replace(',',''))

            #addpro = li_tag.get('data-prdnm')
            
            result_list.append({
                'data-optitemno': optitemno,
                'data-dtloptnm': dtloptnm,
                'data-price': opt_price-base_price,
                #'data-prdnm': addpro
            })

        df = pd.DataFrame(result_list)
        optName_list = df['data-dtloptnm'].to_list()
        str_optName_list = []
        strStock_list = []

        for element in optName_list:
            str_optName_list.append(str(element))
            strStock_list.append('100')

        optPrice_list = df['data-price'].to_list()
        str_optPrice_list = []
        
        for element in optPrice_list:
            str_optPrice_list.append(str(element))

        optName = '|'.join(str_optName_list)
        optPrice = '|'.join(str_optPrice_list)
        stockNum = '|'.join(strStock_list)

    except:
        optName = ""
        optPrice =""
        stockNum = '100'
       
    #상세페이지 추출: 상세페이지는 iframe으로 되어 있어서 해당 iframe으로 전환 후 페이지를 다시 읽어 가져와야 함.

    try:
        driver.switch_to.frame("prdDescIfrm")
        iframe_page = driver.page_source
        soup_desc = BeautifulSoup(iframe_page, "html.parser")
        img_tags = soup_desc.find_all('img')
        #print(f'태그 컬렉션 원본 : {img_tags}')
        
        image_urls=[]
        desc_img = []
        img_tag=''
        
        #url만 하니씩 뽑을 때(다운로드 등..)
        for img_tag in img_tags:
            if 'src' in img_tag.attrs:
                src_url = img_tag['src']
                if 'rapid-up' in src_url:
                    continue
                else:
                    image_urls.append(src_url)
        

        for url in image_urls:
            str_url = '<img src="'+url+'"/>'
            desc_img.append(str_url)
        
        desc_txt = '<br>'.join(desc_img)
        desc_txt = '<p style="text-align: center;"><div>'+desc_txt+'</div></p>'

    except:
        print('아이프레임 못찾음')
    str_count_num = str(count_num)
    driver.switch_to.default_content()
    print("[id: " + id_num + '번 데이터 추출 성공] ' + str_count_num + '/' +url_count +"번째")
    print(Fore.LIGHTBLUE_EX + "<<<<<<<   id: " + id_num + '번성공 ' + str_count_num + '/' +url_count +"개  >>>>>>>" + Fore.RESET, file=f)
    driver.quit()

    return mainImage, addImage1, addImage2, addImage3, title, base_price, optName, optPrice, desc_txt, id_num, stockNum, mainImgList,count_num

# 상세 이미지 url 추출 및 다운로드    
def descImg_Download(descPages,file_path,id_num,num): 
    file_path = file_path
    descimgNum = 0
    id_num = str(id_num)
    mod_urls = []
    
    img_tags = re.findall(r'<img[^>]*src="([^"]+?\.(?:jpg|jpeg|png|gif))', descPages)
    
    if not img_tags:
        img_tags = re.findall(r'https?://[^"]+?\.(?:jpg|jpeg|png|gif)', descPages)

    for url in img_tags: 
        mod_urls.append(url)

    for i in mod_urls:
        file_ext = i.split('.')[-1] # 확장자 추출
        if num == 1:
            path = file_path + '/' + 'desc_' + id_num + '_' + str(descimgNum) + '.' + file_ext

        else:
            path = file_path + '/' + 'main_' + id_num +'_' + str(descimgNum) + '.' + file_ext
    

        random_number = round(random.uniform(0.2, 0.5), 2)
        time.sleep(random_number)
        #path = path_Desc + '/' + id_num + '/' + id_num + '_' + descimgNum + '.' + file_ext
        try:
            #urllib.request.urlretrieve(i, path)
            response = requests.get(i)
            print(f'{descimgNum}번 이미지 다운 성공')

            if response.status_code == 200:
                with open(path, "wb") as file:
                    file.write(response.content)
            else:
                print("이미지 다운로드 실패 Status code:", response.status_code)
                print('오류 있는 '+str(descimgNum)+'번째 이미지 주소: ',i,'\n(url을 콘트롤키+클릭하면 브라우저에서 오픈합니다.)\n')

        except urllib.error.HTTPError:
            print(Fore.RED + '다운로드실패 - 해외쇼핑몰 페이지가 사라졌거나 올바른 상세 url이 아닙니다.'+ Fore.RESET)
            print(str(descimgNum)+'번 오류 이미지주소: ',i,response.status_code)
            continue
        
        except urllib.error.URLError:
            print(Fore.RED + '오류 - 올바른 url이 아닙니다.' + Fore.RESET)
            print('오류 있는 '+str(descimgNum)+'번째 이미지 주소: ',i,'\n(url을 콘트롤키+클릭하면 브라우저에서 오픈합니다.)\n')
            continue
        descimgNum +=1

def make_html(title, desc_txt, top_image, top_image2, bottom_image,bottom_image2): #상세페이지 작성 기능
    descPages = '<div align="center"><!-- 상세페이지 수정은 여기서부터 -->' + desc_txt + '<!-- 상세페이지 수정은 여기까지 --></div>'
    top_img = '<div align="center"><!-- 여기서부터 상단 공지 이미지 --><img src="' + top_image + '"/></div>'
    top2_img = '<br>'+'<div align="center"><img src="' + top_image2 + '"/></div>'
    descTitle = '<br><br><h1 style="text-align: center;"><strong>' + title + "</strong></h1><br><br>"+'<br>'
    bottom_img ='<br>'+ '<div align="center"><!-- 여기서부터 하단 공지1 이미지 --><img src="' + bottom_image + '"/></div>'
    bottom2_img ='<br>'+ '<div align="center"><!-- 여기서부터 하단 공지2 이미지 --><img src="' + bottom_image2 + '"/></div>'
            
    descHtml = top_img + top2_img + descTitle + descPages + bottom_img + bottom2_img
    print("상세페이지 작성 완료!")
    return descHtml

#프로그램 시작
np.set_printoptions(threshold=np.inf, linewidth=np.inf)
init()
print(Fore.LIGHTBLUE_EX + "11번가 상품 수집을 시작합니다. 작성중...")
print(Fore.RESET)


password = loadPassword() #set.ini 파일에서 패스워드를 읽는 함수
#개발자 테스트용 페이지
passTag = getPtag("https://sites.google.com/view/testexmaker/home/11stlogin") #관리자 패스워드가 저장된 웹페이지 url을 전달하여 패스워드를 크롤링 해 오는 getPtag 함수 실행
judge(password,passTag) #웹에서 가져온 패스워드와 set.ini 파일에 저장된 패스워드를 비교하여 틀리면 입력창으로 입력받고 맞으면 통과시킴

urls_path = './urls.xlsx'
urls_DF = readExcel(urls_path)
urls_list = urls_DF['url'].tolist()
url_count = str(len(urls_list))
log_dir = './history'
createFolder(log_dir)
f = open(f'./history/log_{tday_s}.txt', 'w')


print(f'{tday_v} 수집대상 전체 : {url_count}개', file=f)
print(f'수집대상 전체 : {url_count}개')

source_file = './loadingform.xlsx' #11번가 엑셀 서식 불러오기
source_workbook = load_workbook(source_file)
source_sheet = source_workbook['대량등록 양식']

# 기록할 파일 셋팅
target_workbook = Workbook()
target_sheet = target_workbook.active

# 서식파일에서 제목 복사
for row in source_sheet.iter_rows(min_row=1, max_row=2, values_only=True):
    target_sheet.append(row)
    target_sheet.row_dimensions[1].height = 24

load_wb = load_workbook("./setting.xlsx", data_only=True)
load_ws = load_wb['Sheet1']

nickName = str(load_ws.cell(4,1).value)
dropship = str(load_ws.cell(4,2).value)
delivery = str(load_ws.cell(4,3).value)
shipping_template = str(load_ws.cell(4,4).value)
shipping_price_type = str(load_ws.cell(4,5).value)
Shipping_price = str(load_ws.cell(4,6).value)
return_price = str(load_ws.cell(4,7).value)
return_price_free = str(load_ws.cell(4,8).value)
exchange_price = str(load_ws.cell(4,9).value)
after_service_info = str(load_ws.cell(4,10).value)
return_info = str(load_ws.cell(4,11).value)
factory = str(load_ws.cell(4,12).value)
top_image = str(load_ws.cell(4,13).value)
top_image2 = str(load_ws.cell(4,14).value)
bottom_image = str(load_ws.cell(4,15).value)
bottom_image2 = str(load_ws.cell(4,16).value)
shop_name = str(load_ws.cell(4,17).value)
as_info = str(load_ws.cell(4,18).value)
desc_down = str(load_ws.cell(4,19).value)
rand_min = int(load_ws.cell(4,20).value)
rand_max = int(load_ws.cell(4,21).value)

num_row = 3
count_num = 1 

for url in urls_list:
    driver = set_selenium()
    filtered_urls = urls_DF.loc[urls_DF['url'] == url]
    numbering_id = filtered_urls['id']
    id_num = str(numbering_id.iloc[0])
    soup = scrapping(url,driver)
    categori_num = url_parser(soup)
    print('')
    print(url)

    try:
        #크롤링 시작
        mainImage, addImage1, addImage2, addImage3, title, base_price, optName, optPrice, desc_txt, id_num, stockNum, mainImgList,count_num = extractComp(soup, id_num, url_count,count_num)
        # 크롤링 끝

    except:
        num_row += 1
        target_sheet.cell(row = num_row, column = 11).value = f"id:{id_num} 수집실패"
        print(Fore.RED + 'id:'+ id_num+'번 수집실패\n'+Fore.RESET, file=f)
        print(Fore.RED + 'id:'+ id_num+'번 수집실패\n'+Fore.RESET)
        driver.quit()
        continue
    count_num +=1

    rand_price = random.randint(rand_min,rand_max)

    if base_price >= 1000:

        try:
            sell_price = int(base_price)- rand_price
            sell_price = round(sell_price/10)*10
            print(f'판매가격: {sell_price}원', file=f)
            print(f'판매가격: {sell_price}원')
            print(f'랜덤차감: {rand_price}원', file=f)
            print(f'랜덤차감: {rand_price}원')
        except:
            print(Fore.RED+'판매금액을 확인하세요.'+Fore.RESET)
    else:
        print(Fore.RED+'판매금액을 확인하세요.'+Fore.RESET)

    #상세페이지 제작
    descHtml = make_html(title, desc_txt, top_image, top_image2, bottom_image, bottom_image2)

    #이미지 폴더 생성
    mainjpg = ','.join(mainImgList)
    path_Desc = './result/img_desc/' + id_num
    path_Main = './result/img_main'

    try:
        # 이미지 다운로드
        createFolder(path_Main)
        print(Fore.GREEN + '메인이미지 다운로드 중..'+ Fore.RESET)
        descImg_Download(mainjpg,path_Main,id_num,0)
        print(Fore.GREEN + '메인이미지 다운로드완료'+ Fore.RESET)
    except:
        print("메인이미지 실패", file=f)
        print("메인이미지 실패")

    if desc_down == '1':
        try:
            createFolder(path_Desc)
            print(Fore.GREEN + '상세 이미지 다운로드 중..'+ Fore.RESET)    
            descImg_Download(desc_txt,path_Desc,id_num,1)
            print(Fore.GREEN + '상세 이미지 다운로드완료'+ Fore.RESET)
        except:
            print("상세이미지 실패", file=f)
            print("상세이미지 실패")
    else:
        print("상세이미지 다운로드는 패스")
        print("상세이미지 다운로드는 패스", file=f)
    # 엑셀 작성
    target_sheet.cell(row = num_row, column = 1).value = ''
    target_sheet.cell(row = num_row, column = 2).value = nickName
    target_sheet.cell(row = num_row, column = 3).value = 'N'
    target_sheet.cell(row = num_row, column = 4).value = '01'   
    target_sheet.cell(row = num_row, column = 10).value = categori_num
    target_sheet.cell(row = num_row, column = 11).value = title
    target_sheet.cell(row = num_row, column = 15).value = '02'
    target_sheet.cell(row = num_row, column = 16).value = '1287'
    target_sheet.cell(row = num_row, column = 18).value = id_num
    target_sheet.cell(row = num_row, column = 19).value = '01'
    target_sheet.cell(row = num_row, column = 20).value = dropship
    target_sheet.cell(row = num_row, column = 21).value = '01'
    target_sheet.cell(row = num_row, column = 22).value = 'Y'
    target_sheet.cell(row = num_row, column = 23).value = mainImage
    target_sheet.cell(row = num_row, column = 24).value = addImage1
    target_sheet.cell(row = num_row, column = 25).value = addImage2
    target_sheet.cell(row = num_row, column = 26).value = addImage3
    target_sheet.cell(row = num_row, column = 27).value = descHtml
    target_sheet.cell(row = num_row, column = 28).value = 'Y'
    target_sheet.cell(row = num_row, column = 29).value = 'Y'
    target_sheet.cell(row = num_row, column = 30).value = '108'
    target_sheet.cell(row = num_row, column = 32).value = sell_price
    target_sheet.cell(row = num_row, column = 41).value = '01'
    target_sheet.cell(row = num_row, column = 42).value = '선택'
    target_sheet.cell(row = num_row, column = 43).value = optName
    target_sheet.cell(row = num_row, column = 44).value = optPrice
    target_sheet.cell(row = num_row, column = 45).value = stockNum
    target_sheet.cell(row = num_row, column = 54).value = 'N'
    target_sheet.cell(row = num_row, column = 59).value = '02'
    target_sheet.cell(row = num_row, column = 60).value = '01'
    target_sheet.cell(row = num_row, column = 61).value = '00034'
    target_sheet.cell(row = num_row, column = 62).value = shipping_template
    target_sheet.cell(row = num_row, column = 64).value = shipping_price_type
    target_sheet.cell(row = num_row, column = 65).value = Shipping_price
    target_sheet.cell(row = num_row, column = 68).value = '03'
    target_sheet.cell(row = num_row, column = 69).value = return_price
    target_sheet.cell(row = num_row, column = 70).value = return_price_free
    target_sheet.cell(row = num_row, column = 71).value = exchange_price
    target_sheet.cell(row = num_row, column = 72).value = after_service_info
    target_sheet.cell(row = num_row, column = 73).value = return_info
    target_sheet.cell(row = num_row, column = 74).value = 'D'
    target_sheet.cell(row = num_row, column = 76).value = '891045'
    target_sheet.cell(row = num_row, column = 77).value = '23759100'
    target_sheet.cell(row = num_row, column = 78).value = '중국'
    target_sheet.cell(row = num_row, column = 79).value = '23756033'
    target_sheet.cell(row = num_row, column = 80).value = '구매대행 제품으로 해당사항 없음'
    target_sheet.cell(row = num_row, column = 81).value = '11905'
    target_sheet.cell(row = num_row, column = 82).value = shop_name
    target_sheet.cell(row = num_row, column = 83).value = '23760413'
    target_sheet.cell(row = num_row, column = 84).value = as_info
    target_sheet.cell(row = num_row, column = 85).value = '11800'
    target_sheet.cell(row = num_row, column = 86).value = title
    target_sheet.cell(row = num_row, column = 105).value = factory
    target_sheet.cell(row = num_row, column = 115).value = '01|02|02\n02|03\n03|03\n04|05'
    print('수집완료! - 엑셀파일 기록완료')
    num_row += 1
# 크롤링 종료

print('result 파일 저장중')
target_file = f'./result/'+'result_'+tday_s+'.xlsx'
target_workbook.save(target_file)
print('result 파일 저장완료!', file=f)
print('result 파일 저장완료!')
print(Fore.YELLOW + "모든 작업 완료. 엔터를 누르면 종료합니다."+Fore.RESET)
f.close()
aInput = input("")
sys.exit()  
